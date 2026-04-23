"""
Smarket TMS — Optimizador de Ruteo
Desarrollado con Streamlit
"""

import streamlit as st
import pandas as pd
import math
import json
import io
from typing import Optional

try:
    import gspread
    from google.oauth2.service_account import Credentials
    GSHEETS_OK = True
except ImportError:
    GSHEETS_OK = False

# ══════════════════════════════════════════════════════════
# CONFIG
# ══════════════════════════════════════════════════════════
st.set_page_config(
    page_title="Smarket TMS",
    page_icon="🚛",
    layout="wide",
    initial_sidebar_state="expanded",
)

st.markdown("""
<style>
.ruta-card {
    background: var(--background-color, #fff);
    border: 1px solid #e5e7eb;
    border-radius: 12px;
    padding: 16px;
    margin-bottom: 12px;
    box-shadow: 0 1px 4px rgba(0,0,0,.06);
}
.ruta-header {
    display: flex;
    justify-content: space-between;
    align-items: center;
    margin-bottom: 10px;
    flex-wrap: wrap;
    gap: 6px;
}
.ruta-title { font-size: 15px; font-weight: 600; }
.badge {
    display: inline-block;
    padding: 3px 10px;
    border-radius: 99px;
    font-size: 11px;
    font-weight: 600;
}
.badge-green  { background:#dcfce7; color:#166534; }
.badge-yellow { background:#fef9c3; color:#854d0e; }
.badge-red    { background:#fee2e2; color:#991b1b; }
.badge-blue   { background:#dbeafe; color:#1e40af; }
.badge-gray   { background:#f3f4f6; color:#374151; }
.stat-row {
    display: flex;
    gap: 16px;
    font-size: 12px;
    color: #6b7280;
    margin-bottom: 8px;
    flex-wrap: wrap;
}
.stat-val { font-weight: 600; color: #111; }
.ped-row {
    display: flex;
    justify-content: space-between;
    align-items: flex-start;
    padding: 6px 0;
    border-top: 1px solid #f3f4f6;
    font-size: 12px;
    gap: 8px;
}
.ped-imp-ok   { color:#166534; font-weight:700; }
.ped-imp-warn { color:#b45309; font-weight:700; }
.ped-imp-bad  { color:#991b1b; font-weight:700; }
.progress-wrap {
    background:#f3f4f6;
    border-radius:4px;
    height:6px;
    overflow:hidden;
    margin-top:4px;
}
.progress-fill { height:100%; border-radius:4px; }
</style>
""", unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════
# DATOS DE MUESTRA
# ══════════════════════════════════════════════════════════
SAMPLE_PEDIDOS = pd.DataFrame([
    {"id":"35268","cliente":"Sebastian Duarte (1)","direccion":"Av. S. Martin 7300","zona":"Villa Devoto","localidad":"Villa Devoto","cp":1419,"kilos":114,"valor":570137,"bultos":74},
    {"id":"35269","cliente":"Sebastian Duarte (2)","direccion":"Av. S. Martin 7300","zona":"Villa Devoto","localidad":"Villa Devoto","cp":1419,"kilos":5,"valor":60077,"bultos":1},
    {"id":"35250","cliente":"Buenos Alimentos BSM1644","direccion":"Av. General Paz 12301","zona":"Lomas del Mirador","localidad":"Lomas del Mirador","cp":1752,"kilos":125,"valor":755344,"bultos":10},
    {"id":"35178","cliente":"Servicios en Gastronomia SRL","direccion":"Lisandro de la Torre 1609","zona":"Lomas del Mirador","localidad":"Lomas del Mirador","cp":1752,"kilos":55,"valor":507580,"bultos":6},
    {"id":"35241","cliente":"Outletbar Misiones","direccion":"Perdriel 1151","zona":"CABA","localidad":"CABA","cp":1279,"kilos":48,"valor":224813,"bultos":28},
    {"id":"35265","cliente":"Fuerzas del Interior SRL","direccion":"Av. Garmendia 4813","zona":"Chacarita","localidad":"Chacarita","cp":1427,"kilos":175,"valor":388527,"bultos":13},
    {"id":"35282","cliente":"Cincunegui Jorge","direccion":"Montevideo 274","zona":"Tigre","localidad":"Tigre","cp":1648,"kilos":27,"valor":180113,"bultos":3},
    {"id":"35141","cliente":"Francisco Orlando Chavez","direccion":"Cordero 3821","zona":"San Fernando","localidad":"San Fernando","cp":1645,"kilos":782,"valor":4176564,"bultos":64},
    {"id":"35232","cliente":"Dam Eventos S.R.L.","direccion":"Eduardo Wilde 1561","zona":"Francisco Álvarez","localidad":"Francisco Álvarez","cp":1746,"kilos":289,"valor":1371475,"bultos":34},
    {"id":"35188","cliente":"Ramirez Gustavo","direccion":"San Martin esq. Italia","zona":"Luján","localidad":"Luján","cp":6700,"kilos":118,"valor":1049708,"bultos":12},
])

# Lista de transportistas "únicos" (solo pueden hacer 1 recorrido por día)
TRANSPORTISTAS_UNICOS = {"Gaby", "Juan", "Ezequiel"}

SAMPLE_TRANSPORTES = pd.DataFrame([
    {"nombre":"Gaby","capacidad_kg":600,"costo_hora":22000,"costo_fijo":0,"ayudante":False},
    {"nombre":"Juan","capacidad_kg":600,"costo_hora":22000,"costo_fijo":0,"ayudante":False},
    {"nombre":"Ezequiel","capacidad_kg":600,"costo_hora":22000,"costo_fijo":0,"ayudante":False},
    {"nombre":"Greco - Utilitario","capacidad_kg":600,"costo_hora":18000,"costo_fijo":0,"ayudante":False},
    {"nombre":"Greco - Sprinter","capacidad_kg":2800,"costo_hora":35000,"costo_fijo":50000,"ayudante":True},
    {"nombre":"Greco - MB 608","capacidad_kg":3500,"costo_hora":48000,"costo_fijo":50000,"ayudante":True},
    {"nombre":"Greco - MB 1114","capacidad_kg":10000,"costo_hora":60000,"costo_fijo":0,"ayudante":False},
])

# ══════════════════════════════════════════════════════════
# CORREDORES DE RUTEO
# ══════════════════════════════════════════════════════════
# Cada corredor define las localidades que se pueden visitar en un SOLO viaje.
# El motor elige el corredor MÁS CHICO que cubra todos los pedidos a rutear,
# así prioriza viajes concentrados geográficamente y llena los vehículos.

def _norm_loc(s):
    """Normaliza una localidad: minúsculas, sin tildes, espacios simples."""
    import unicodedata
    s = str(s).strip().lower()
    s = unicodedata.normalize("NFD", s)
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    # Quitar guiones y espacios dobles
    s = " ".join(s.replace("-", " ").split())
    return s

CORREDORES = {
    "Norte 1": [
        "San Isidro","Beccar","San Fernando","Martínez","Villa Adelina","Acassuso",
        "Tigre","Rincón de Milberg","Pacheco","Benavídez","Belén de Escobar",
        "Dique Luján","Garín",
    ],
    "Norte 2": [
        "San Isidro","Beccar","San Fernando","Martínez","Acassuso","Villa Adelina",
        "Tigre","Pacheco","Garín","Don Torcuato","Grand Bourg","Tortuguitas",
        "Del Viso","Manuel Alberti","Pilar","Fátima","Zelaya","Exaltación de la Cruz",
    ],
    "Noroeste 1": [
        "San Isidro","Beccar","San Fernando","Tigre","Martínez","Acassuso","Munro",
        "Villa Adelina","Villa Ballester","Villa José León Suárez","Villa Bonich",
        "Hurlingham","William C. Morris","Bella Vista","San Miguel","Moreno",
        "Ituzaingó","Paso del Rey","Merlo","Francisco Álvarez","General Rodríguez","Luján",
    ],
    "Noroeste 2": [
        "San Isidro","Beccar","San Fernando","Tigre","Martínez","Munro","Acassuso",
        "La Lucila","Olivos","Florida","Vicente López","Villa Urquiza","Villa Maipú",
        "San Martín","Villa Devoto","Caseros","Liniers","Ramos Mejía","Haedo",
        "Villa Tesei","Morón","Castelar","Ituzaingó","Merlo","Paso del Rey","Moreno",
        "Francisco Álvarez","General Rodríguez","Luján",
    ],
    "Oeste": [
        "San Isidro","Beccar","San Fernando","Tigre","Martínez","Munro","Acassuso",
        "La Lucila","Olivos","Florida","Vicente López","Villa Urquiza","Villa Maipú",
        "San Martín","Villa Devoto","Caseros","Liniers","Ramos Mejía","Parque Avellaneda",
        "Villa Lugano","Tapiales","La Tablada","San Justo","Aldo Bonzi",
        "Isidro Casanova","Gregorio de Laferrere","González Catán","Virrey del Pino",
    ],
    "Norte-CABA-Sur": [
        "San Isidro","Beccar","San Fernando","Tigre","Martínez","Munro","Acassuso",
        "La Lucila","Olivos","Florida","Vicente López","Villa Urquiza","Villa Maipú",
        "San Martín","Villa Devoto","Caseros","Liniers","Flores","Parque Avellaneda",
        "Villa Lugano","Tapiales","Lomas de Zamora","Burzaco","Monte Grande","Ezeiza",
        "Lanús","Adrogué","Temperley","Avellaneda",
    ],
    "Sur 1": [
        "San Isidro","Beccar","San Fernando","Tigre","Martínez","Munro","Acassuso",
        "La Lucila","Olivos","Florida","Vicente López","Villa Urquiza","Núñez",
        "Belgrano","Palermo","Recoleta","CABA","Monserrat","San Telmo","Barracas",
        "Villa Soldati","Puerto Madero","La Boca","Parque Patricios","Nueva Pompeya",
        "Chacarita","Caballito","Floresta","Colegiales","Villa Crespo","Almagro","San Cristóbal",
        "Avellaneda","Sarandí","Wilde","Bernal","Quilmes","Berazategui","Ranelagh",
        "Hudson","Florencio Varela","Villa Elisa","Gonnet","La Plata","Berisso",
        "Ensenada","Villa Elvira","Lisandro Olmos",
        # Barrios CABA norte/oeste que geográficamente acompañan este recorrido
        "La Paternal","Villa del Parque","Villa Pueyrredón","Boedo","San Cristóbal",
        "Parque Chacabuco","Flores","Floresta","Mataderos","Villa Lugano",
    ],
    "CABA": [
        "San Isidro","Beccar","San Fernando","Tigre","Martínez","Munro","Acassuso",
        "La Lucila","Olivos","Florida","Vicente López","Villa Urquiza","Núñez",
        "Belgrano","Palermo","Recoleta","CABA","Monserrat","San Telmo","Barracas",
        "Villa Soldati","Puerto Madero","La Boca","Parque Patricios","Nueva Pompeya",
        "Chacarita","Caballito","Floresta","Colegiales","Villa Crespo","Almagro","San Cristóbal",
        "La Paternal","Villa del Parque","Villa Pueyrredón","Boedo","Parque Chacabuco",
        "Flores","Mataderos","Villa Lugano","Liniers","Devoto","Villa Devoto",
    ],
}

# Tamaño de cada corredor (para preferir los más chicos cuando son equivalentes)
CORREDORES_SIZE = {k: len(v) for k, v in CORREDORES.items()}

# Índice: localidad_normalizada → set de corredores que la incluyen
_LOC_TO_CORREDORES = {}
for corr, locs in CORREDORES.items():
    for l in locs:
        key = _norm_loc(l)
        _LOC_TO_CORREDORES.setdefault(key, set()).add(corr)

# Alias para localidades con nombres alternativos frecuentes
_ALIASES = {
    "capital federal": "caba",
    "ciudad autonoma de buenos aires": "caba",
    "ciudad autónoma de buenos aires": "caba",
    "lomas del mirador": "lanus",   # aproximación
    "villa luro": "villa devoto",   # cercanía
    "villa luro devoto": "villa devoto",
    "villa bosch": "villa ballester",  # colindante
    "la paternal": "caba",
    "villa del parque": "caba",
    "villa pueyrredon": "caba",
    "villa santa rita": "caba",
    "mataderos": "caba",
    "boedo": "caba",
    "saavedra": "caba",
    "jose c paz": "san miguel",
    "malvinas argentinas": "grand bourg",
    "san isidro": "san isidro",
    "beccar": "beccar",
    "béccar": "beccar",
}

def corredores_de_localidad(loc_str):
    """Devuelve el set de corredores que contienen a esta localidad."""
    key = _norm_loc(loc_str)
    # 1. Match directo
    if key in _LOC_TO_CORREDORES:
        return _LOC_TO_CORREDORES[key]
    # 2. Alias
    if key in _ALIASES:
        aliased = _ALIASES[key]
        if aliased in _LOC_TO_CORREDORES:
            return _LOC_TO_CORREDORES[aliased]
    # 3. Match por prefijo/substring
    for k, corrs in _LOC_TO_CORREDORES.items():
        if key and (key in k or k in key):
            return corrs
    # 4. Fallback: CABA (todos los corredores excepto Norte puros)
    return {"CABA"}

# Retrocompatibilidad: ZONAS_TRANS define el transportista "natural" por nombre de zona
# (solo usado cuando el usuario carga pedidos con "zona" ya asignada, no localidad)
ZONAS_TRANS = {
    "Norte 1":        "Greco - Sprinter",
    "Norte 2":        "Greco - Sprinter",
    "Noroeste 1":     "Greco - Sprinter",
    "Noroeste 2":     "Greco - Sprinter",
    "Oeste":          "Greco - Sprinter",
    "Norte-CABA-Sur": "Greco - Sprinter",
    "Sur 1":          "Greco - Sprinter",
    "CABA":           "Juan",
    # Retrocompatibilidad con zonas viejas
    "Sur-Oeste":    "Gaby",
    "Norte":        "Greco - Sprinter",
    "Acceso Oeste": "Greco - Sprinter",
}

AYUDANTE_COSTO = 50000
MAX_PARADAS    = 8

# Flota Greco: cada entrada es UN vehículo disponible por día (1 chofer = 1 salida)
# Orden: del más grande al más chico, para asignar primero el vehículo más eficiente
FLOTA_GRECO = [
    "Greco - Sprinter",    # chofer 1 → 2800kg
    "Greco - Utilitario",  # chofer 2 → 600kg (Kangoo)
    # MB 608 y MB 1114 se usan si los anteriores no alcanzan o se agregan manualmente
]

# ══════════════════════════════════════════════════════════
# UTILIDADES
# ══════════════════════════════════════════════════════════
def ic_badge(v):
    if v <= 3: return "badge-green"
    if v <= 5: return "badge-yellow"
    return "badge-red"

def ic_row(v):
    if v <= 3: return "ped-imp-ok"
    if v <= 5: return "ped-imp-warn"
    return "ped-imp-bad"

def carga_col(p):
    if p >= 75: return "#16a34a"
    if p >= 45: return "#d97706"
    return "#dc2626"

def imp_ped(kg_p, val_p, kg_t, flete):
    if kg_t <= 0 or val_p <= 0: return 0.0
    return (flete * kg_p / kg_t) / val_p * 100

def auto_horas(n_paradas):
    h = 80 / 35 + n_paradas * 20 / 60
    return math.ceil(h * 2) / 2

def auto_flete(horas, tarifa, fijo, ayudante):
    return horas * tarifa + fijo + (AYUDANTE_COSTO if ayudante else 0)

# ══════════════════════════════════════════════════════════
# NORMALIZAR COLUMNAS  ← FIX PRINCIPAL
# ══════════════════════════════════════════════════════════
def normalizar_columnas(df: pd.DataFrame) -> Optional[pd.DataFrame]:
    """
    Paso 1 — convierte TODAS las columnas a minusculas (fix para Zona, ZONA, etc.)
    Paso 2 — renombra variantes conocidas al esquema interno
    Paso 3 — si no encuentra 'zona' muestra error amigable y retorna None
    Pasos 4-7 — valores por defecto, parseo numerico, filtros
    """

    # ── Paso 1: minusculas ────────────────────────────────
    df.columns = [str(c).strip().lower() for c in df.columns]

    # ── Paso 2: renombrar variantes ───────────────────────
    MAP = {
        # id
        "n° de pedido": "id", "n° pedido": "id", "nro pedido": "id",
        "pedido": "id", "n_pedido": "id", "id pedido": "id",
        "numero de pedido": "id", "num pedido": "id",
        # cliente
        "descripcion cliente": "cliente",
        "descripción cliente": "cliente",
        "descripcion_cliente": "cliente",
        "nombre cliente": "cliente",
        "razon social": "cliente",
        "razón social": "cliente",
        # direccion
        "direccion de entrega": "direccion",
        "dirección de entrega": "direccion",
        "direccion_entrega": "direccion",
        "dirección": "direccion",
        "domicilio": "direccion",
        "domicilio de entrega": "direccion",
        # zona — todas las variantes posibles
        "zona": "zona",
        "localidad": "zona",
        "localidad de entrega": "zona",
        "localidad_de_entrega": "zona",
        "barrio/expreso": "zona",
        "barrio": "zona",
        "ciudad": "zona",
        "region": "zona",
        "región": "zona",
        "partido": "zona",
        "municipio": "zona",
        # cp
        "cp": "cp",
        "codigo postal": "cp",
        "código postal": "cp",
        "cod postal": "cp",
        "cod. postal": "cp",
        "cód. postal": "cp",
        "postal": "cp",
        # kilos
        "kilos": "kilos",
        "kg": "kilos",
        "peso": "kilos",
        "peso kg": "kilos",
        "peso (kg)": "kilos",
        "kilos netos": "kilos",
        "peso neto": "kilos",
        # valor
        "valor": "valor",
        "importe del pedido": "valor",
        "valor venta": "valor",
        "importe": "valor",
        "monto": "valor",
        "total": "valor",
        "precio": "valor",
        # bultos
        "bultos": "bultos",
        "cajas": "bultos",
        "cant. bultos": "bultos",
        "cantidad bultos": "bultos",
        "unidades": "bultos",
    }
    df = df.rename(columns={k: v for k, v in MAP.items() if k in df.columns})

    # ── Paso 2.5: deduplicar columnas (si hay dos que mapean al mismo nombre) ──
    # Ej: si el archivo tenia "Localidad" y "Localidad de entrega", ambas quedan "zona".
    # Nos quedamos con la primera no-vacia.
    if df.columns.duplicated().any():
        # Para cada nombre duplicado, tomar la primera columna que tenga datos
        new_cols = {}
        for col_name in df.columns.unique():
            cols_with_name = [c for i, c in enumerate(df.columns) if c == col_name]
            if len([c for c in df.columns if c == col_name]) > 1:
                # Hay duplicados: tomar el sub-DataFrame y combinar
                sub = df.loc[:, df.columns == col_name]
                # Priorizar la primera columna con datos no nulos
                combined = sub.iloc[:, 0]
                for i in range(1, sub.shape[1]):
                    combined = combined.fillna(sub.iloc[:, i])
                    # Si la primera columna esta vacia en una fila, usar la siguiente
                    mask_empty = combined.isna() | (combined.astype(str).str.strip() == "")
                    combined.loc[mask_empty] = sub.iloc[:, i][mask_empty]
                new_cols[col_name] = combined
            else:
                new_cols[col_name] = df[col_name]
        df = pd.DataFrame(new_cols)

    # ── Paso 3: validar columna critica ───────────────────
    if "zona" not in df.columns:
        cols_encontradas = ", ".join(f"\'{c}\'" for c in df.columns.tolist())
        st.error(
            "❌ **Error: No se encontró la columna `zona` en el archivo.**\n\n"
            "Verificá que tu archivo tenga una columna llamada: "
            "`zona`, `localidad`, `localidad de entrega`, `barrio`, `ciudad`, `región` o `partido`.\n\n"
            f"**Columnas encontradas en tu archivo:** {cols_encontradas}"
        )
        return None

    # ── Paso 4: columnas numericas opcionales ─────────────
    for col, default in [("cp", 0), ("kilos", 0), ("valor", 0), ("bultos", 0)]:
        if col not in df.columns:
            df[col] = default

    # ── Paso 5: columnas de texto opcionales ──────────────
    for col in ("id", "cliente", "direccion"):
        if col not in df.columns:
            df[col] = ""

    # ── Paso 6: parseo numerico seguro ────────────────────
    df["kilos"] = pd.to_numeric(df["kilos"], errors="coerce").fillna(0)
    df["valor"] = pd.to_numeric(df["valor"], errors="coerce").fillna(0)
    df["cp"]    = pd.to_numeric(df["cp"],    errors="coerce").fillna(0).astype(int)

    # ── Paso 7: filtrar deposito y filas sin kilos ────────
    mask_dep = df["direccion"].str.lower().str.contains("maestro santana|beccar", na=False)
    df = df[~mask_dep]
    df = df[df["kilos"] > 0]

    if df.empty:
        st.warning("⚠️ No quedaron pedidos después de filtrar. Verificá que el archivo tenga filas con kilos > 0.")
        return None

    # ── Paso 8: preservar localidad real + calcular corredores compatibles ──
    # No pisamos la zona original porque la queremos mostrar en las tarjetas.
    # En vez de eso, guardamos para cada pedido los corredores en los que "cabe".
    df["localidad"] = df["zona"].astype(str)
    df["_corredores_ok"] = df["zona"].apply(lambda z: corredores_de_localidad(z))

    return df.reset_index(drop=True)

# ══════════════════════════════════════════════════════════
# MOTOR DE RUTEO  ← GUARD CLAUSE CONTRA None
# ══════════════════════════════════════════════════════════
def _trans_row(nombre, transportes_df):
    """Devuelve la fila del transportista por nombre."""
    m = transportes_df[transportes_df["nombre"] == nombre]
    return m.iloc[0] if not m.empty else transportes_df.iloc[0]


def _elegir_transportista(kg_zona, zona, transportes_df):
    """Devuelve la fila del transportista más adecuado para la zona y el peso."""
    trans_nom = ZONAS_TRANS.get(zona, "Greco - Utilitario")
    match = transportes_df[transportes_df["nombre"] == trans_nom]
    if match.empty:
        greco = transportes_df[transportes_df["nombre"].str.startswith("Greco")]
        greco = greco[greco["capacidad_kg"] >= kg_zona].sort_values("capacidad_kg")
        return greco.iloc[0] if not greco.empty else transportes_df.iloc[0]
    trans_row = match.iloc[0]
    if trans_nom.startswith("Greco") and kg_zona > float(trans_row["capacidad_kg"]):
        greco = transportes_df[transportes_df["nombre"].str.startswith("Greco")]
        greco = greco[greco["capacidad_kg"] >= kg_zona].sort_values("capacidad_kg")
        if not greco.empty:
            trans_row = greco.iloc[0]
    return trans_row


def _elegir_corredor_para_pedidos(peds_df):
    """
    Dado un subconjunto de pedidos:
    1. Si hay un corredor que cubre a TODOS → retorna el más compacto.
    2. Si no → retorna el corredor que cubre a MÁS pedidos (mejor aproximación).
    """
    if peds_df.empty:
        return None

    # Intentar intersección perfecta
    comunes = None
    for _, p in peds_df.iterrows():
        corrs = p.get("_corredores_ok", set()) or set()
        if not isinstance(corrs, (set, frozenset)):
            corrs = set(corrs) if corrs else set()
        if comunes is None:
            comunes = set(corrs)
        else:
            comunes &= corrs

    if comunes:
        return min(comunes, key=lambda c: CORREDORES_SIZE.get(c, 999))

    # Sin corredor común: votar por el que cubre más pedidos
    conteo = {}
    for _, p in peds_df.iterrows():
        corrs = p.get("_corredores_ok", set()) or set()
        for c in corrs:
            conteo[c] = conteo.get(c, 0) + 1

    if not conteo:
        return "CABA"

    max_cov = max(conteo.values())
    # Entre los que cubren más, el más compacto
    candidatos = [c for c, v in conteo.items() if v == max_cov]
    return min(candidatos, key=lambda c: CORREDORES_SIZE.get(c, 999))


def _trans_pref_para_corredor(corredor):
    """Devuelve el transportista sugerido según el corredor."""
    return ZONAS_TRANS.get(corredor, "Greco - Sprinter")


def _bin_packing_por_corredor(peds_df, capacidad):
    """
    Agrupa pedidos en 'bins' del tamaño `capacidad` respetando MAX_PARADAS,
    intentando que todos los pedidos de un bin compartan al menos un corredor.

    Estrategia:
    1. Ordenar por kilos desc (first-fit decreasing).
    2. Para cada pedido, buscar un bin existente donde:
       - quede capacidad suficiente,
       - quede espacio de paradas,
       - la intersección de corredores siga siendo no vacía.
    3. Si no entra en ninguno, abrir un bin nuevo.
    Devuelve lista de (lista_indices, corredor_elegido_o_None).
    """
    bins = []  # cada bin: {"ids": [], "kg": 0, "corrs": set}
    for _, p in peds_df.sort_values("kilos", ascending=False).iterrows():
        pid = p.name
        corrs_p = set(p.get("_corredores_ok") or set())
        asignado = False
        for b in bins:
            if b["kg"] + p["kilos"] > capacidad:      continue
            if len(b["ids"]) >= MAX_PARADAS:          continue
            nueva_inter = b["corrs"] & corrs_p
            if not nueva_inter:                       continue
            # Ok, cabe acá
            b["ids"].append(pid)
            b["kg"] += p["kilos"]
            b["corrs"] = nueva_inter
            asignado = True
            break
        if not asignado:
            bins.append({"ids": [pid], "kg": p["kilos"], "corrs": set(corrs_p)})
    # Elegir corredor final para cada bin (el más chico compatible)
    resultado = []
    for b in bins:
        corr = min(b["corrs"], key=lambda c: CORREDORES_SIZE.get(c, 999)) if b["corrs"] else None
        resultado.append((b["ids"], corr))
    return resultado



# ────────────────────────────────────────────────────────
# MOTOR DE RUTEO — FILOSOFÍA:
#   1. Usar la MENOR cantidad de fletes posible.
#   2. Cada flete debe llenarse al menos al 75% de capacidad.
#   3. Si no hay suficiente carga para llenar un vehículo al 75%,
#      absorber esos pedidos en otro flete ya existente.
#   4. Los corredores son una GUÍA geográfica, no una restricción estricta.
#      Si mezclar zonas da un recorrido lógico, se hace.
#   5. Flota Greco: cada vehículo de FLOTA_GRECO sale máximo 1 vez por día.
#      MB 608 / MB 1114 se usan como extras solo si es necesario.
#   6. JG (Gaby, Juan, Ezequiel): 1 salida por chofer, 600kg máx.
# ────────────────────────────────────────────────────────

CARGA_MINIMA_PCT = 0.75   # objetivo mínimo de llenado por vehículo


def _flota_ordenada(transportes_df):
    """
    Devuelve la lista de vehículos disponibles ordenada de mayor a menor capacidad.
    Respeta FLOTA_GRECO (máx 1 salida c/u) y los JG (máx 1 salida c/u).
    """
    vehiculos = []
    # Greco (orden FLOTA_GRECO)
    for nom in FLOTA_GRECO:
        m = transportes_df[transportes_df["nombre"] == nom]
        if not m.empty:
            vehiculos.append({
                "nombre": nom,
                "cap": float(m.iloc[0]["capacidad_kg"]),
                "tarifa": float(m.iloc[0]["costo_hora"]),
                "fijo":   float(m.iloc[0].get("costo_fijo", 0)),
                "ayudante": bool(m.iloc[0].get("ayudante", False)),
                "tipo": "greco_flota",
                "max_salidas": 1,
            })
    # JG
    for nom in ["Gaby", "Juan", "Ezequiel"]:
        m = transportes_df[transportes_df["nombre"] == nom]
        if not m.empty:
            vehiculos.append({
                "nombre": nom,
                "cap": float(m.iloc[0]["capacidad_kg"]),
                "tarifa": float(m.iloc[0]["costo_hora"]),
                "fijo":   float(m.iloc[0].get("costo_fijo", 0)),
                "ayudante": bool(m.iloc[0].get("ayudante", False)),
                "tipo": "jg",
                "max_salidas": 1,
            })
    # Extras (MB 608, MB 1114 — sin límite de salidas)
    for _, row in transportes_df.iterrows():
        nom = row["nombre"]
        if nom in [v["nombre"] for v in vehiculos]: continue
        if not nom.startswith("Greco"): continue
        vehiculos.append({
            "nombre": nom,
            "cap": float(row["capacidad_kg"]),
            "tarifa": float(row["costo_hora"]),
            "fijo":   float(row.get("costo_fijo", 0)),
            "ayudante": bool(row.get("ayudante", False)),
            "tipo": "greco_extra",
            "max_salidas": 99,
        })
    return vehiculos


def _consolidar_zonas(pedidos_df, transportes_df):
    """
    Asigna pedidos a vehículos con la filosofía descripta arriba.
    Devuelve lista de tuplas (corredor, df_pedidos, trans_preferido).
    """
    df = pedidos_df.copy()
    if "_corredores_ok" not in df.columns:
        df["_corredores_ok"] = df["zona"].apply(lambda z: corredores_de_localidad(z))

    # ── Atajo: todo en 1 JG ──────────────────────────────
    kg_total = df["kilos"].sum()
    n_total  = len(df)
    if kg_total <= 600 and n_total <= MAX_PARADAS:
        corr = _elegir_corredor_para_pedidos(df) or "CABA"
        return [(corr, df.drop(columns=["_corredores_ok"], errors="ignore"), "Juan")]

    # ── Preparar flota ───────────────────────────────────
    flota = _flota_ordenada(transportes_df)
    # "bins": slots de carga por vehículo  (puede haber varios slots por extra)
    bins  = []  # {veh: dict, ids: [], kg: 0.0}

    def _nuevo_bin(veh):
        return {"veh": veh, "ids": [], "kg": 0.0}

    # Iniciar un bin por cada vehículo (uno a la vez hasta max_salidas)
    salidas_usadas = {}
    for v in flota:
        salidas_usadas[v["nombre"]] = 0
        bins.append(_nuevo_bin(v))

    # ── Pre-agrupado por dirección idéntica ──────────────
    # Pedidos con la MISMA dirección física deben ir SIEMPRE en el mismo bin.
    # Los agrupamos primero y los tratamos como una unidad inseparable.

    def _norm_dir(d):
        """Normaliza dirección para comparar: minúsculas, sin comas extras."""
        import re
        d = str(d).lower().strip()
        # Quitar provincia, país
        for t in [", buenos aires, argentina", ", ciudad autonoma de buenos aires, argentina",
                  ", argentina", ", buenos aires"]:
            d = d.replace(t, "")
        # Quitar espacios y comas redundantes
        d = re.sub(r"[, ]+", " ", d).strip()
        return d

    # Construir grupos por dirección normalizada
    dir_groups = {}  # {dir_norm: [idx_df, ...]}
    for idx, row in df.iterrows():
        key = _norm_dir(row["direccion"])
        dir_groups.setdefault(key, []).append(idx)

    # Crear "super-pedidos" por grupo (se asignan juntos o no se asignan)
    units = []  # cada unit: [lista de idx_df], kg_total, tipo="group"|"single"
    for key, idxs in dir_groups.items():
        kg = df.loc[idxs, "kilos"].sum()
        units.append({"idxs": idxs, "kg": kg})

    # Ordenar units por kg desc (first-fit decreasing)
    units.sort(key=lambda u: -u["kg"])

    def _puede_entrar(b, kg_p):
        cap = b["veh"]["cap"]
        return (b["kg"] + kg_p <= cap and
                len(b["ids"]) + 1 <= MAX_PARADAS)  # al menos 1 slot libre

    def _puede_entrar_grupo(b, kg_total, n_peds):
        cap = b["veh"]["cap"]
        return (b["kg"] + kg_total <= cap and
                len(b["ids"]) + n_peds <= MAX_PARADAS)

    sobrantes_idx = []
    for unit in units:
        idxs   = unit["idxs"]
        kg_u   = unit["kg"]
        n_u    = len(idxs)
        colocado = False

        # Buscar bin que acepte TODO el grupo junto
        candidatos = [b for b in bins if _puede_entrar_grupo(b, kg_u, n_u)]
        candidatos.sort(key=lambda b: -b["kg"])  # más lleno primero

        for b in candidatos:
            b["ids"].extend(idxs)
            b["kg"] += kg_u
            colocado = True
            break

        if not colocado:
            # Abrir bin extra si existe
            for v in [v for v in flota if v["tipo"] == "greco_extra"]:
                if v["cap"] >= kg_u and n_u <= MAX_PARADAS:
                    nb = _nuevo_bin(v)
                    nb["ids"].extend(idxs)
                    nb["kg"] += kg_u
                    bins.append(nb)
                    colocado = True
                    break
            if not colocado:
                sobrantes_idx.extend(idxs)
            if not colocado:
                sobrantes_idx.append(pid)

    # ── Consolidar: eliminar bins que no lleguen al 75% ──
    # Si un bin está bajo el umbral, redistribuir sus pedidos en otros bins
    # Iterar hasta estabilidad
    cambio = True
    while cambio:
        cambio = False
        bins_activos = [b for b in bins if b["ids"]]
        bins_activos.sort(key=lambda b: b["kg"])  # el más vacío primero

        for b_vacio in bins_activos:
            pct = b_vacio["kg"] / b_vacio["veh"]["cap"]
            if pct >= CARGA_MINIMA_PCT:
                continue  # ya bien cargado, no tocar
            # Intentar mover todos sus pedidos a otros bins
            peds_del_bin = list(b_vacio["ids"])  # copia
            otros = [ob for ob in bins if ob is not b_vacio and ob["ids"]]
            otros.sort(key=lambda ob: -ob["kg"])  # los más llenos primero

            movidos = []
            for pid in sorted(peds_del_bin, key=lambda x: -df.loc[x, "kilos"]):
                kg_p = df.loc[pid, "kilos"]
                for ob in otros:
                    if _puede_entrar(ob, kg_p):
                        ob["ids"].append(pid)
                        ob["kg"] += kg_p
                        movidos.append(pid)
                        break

            if len(movidos) == len(peds_del_bin):
                # Todos se pudieron mover → vaciar este bin
                b_vacio["ids"] = []
                b_vacio["kg"]  = 0.0
                cambio = True
                break  # reiniciar el while

    # ── Sobrantes → último recurso ────────────────────────
    if sobrantes_idx:
        kg_sob = sum(df.loc[pid, "kilos"] for pid in sobrantes_idx)
        greco_rows = transportes_df[transportes_df["nombre"].str.startswith("Greco")]
        greco_rows = greco_rows.sort_values("capacidad_kg")
        cand = greco_rows[greco_rows["capacidad_kg"] >= kg_sob]
        trans = str(cand.iloc[0]["nombre"]) if not cand.empty else "Greco - MB 1114"
        df_sob = df.loc[sobrantes_idx]
        corr_s = _elegir_corredor_para_pedidos(df_sob) or "Varios"
        yield_extra = [(corr_s, df_sob.drop(columns=["_corredores_ok"], errors="ignore"), trans)]
    else:
        yield_extra = []

    # ── Emitir rutas ─────────────────────────────────────
    resultado = []
    for b in bins:
        if not b["ids"]: continue
        df_r = df.loc[b["ids"]].sort_values("cp")
        corr = _elegir_corredor_para_pedidos(df_r) or "Varios"
        resultado.append((corr, df_r.drop(columns=["_corredores_ok"], errors="ignore"), b["veh"]["nombre"]))

    resultado.extend(yield_extra)
    return resultado


def optimizar_rutas(pedidos_df, transportes_df, fletes_man, horas_man, asign_man=None, rutas_extra=None):
    """
    asign_man  : dict {pedido_id: ruta_idx} — asignaciones manuales de pedidos a rutas
    rutas_extra: list — rutas extra vacías. Cada elemento puede ser:
                 - int (índice de ruta, usa transportista default), o
                 - dict {"idx": int, "transportista": str} (con transportista específico)
    """
    # ── Guard: validar entrada ────────────────────────────
    if pedidos_df is None or not isinstance(pedidos_df, pd.DataFrame) or pedidos_df.empty:
        return []
    zona_col = pedidos_df["zona"]
    if isinstance(zona_col, pd.DataFrame):
        zona_col = zona_col.iloc[:, 0]
        pedidos_df = pedidos_df.copy()
        pedidos_df["zona"] = zona_col
    if "zona" not in pedidos_df.columns:
        st.error("❌ **Error: No se encontró la columna `zona` en el archivo.**")
        return []

    asign_man = asign_man or {}
    rutas_extra = rutas_extra or []

    if True:  # siempre ruteo automático primero, asign_man se aplica como post-proceso
        # _consolidar_zonas ya devuelve tuplas de 3 (nombre, df, trans_pref)
        grupos = _consolidar_zonas(pedidos_df, transportes_df)
        # Agregar rutas extra vacías
        for extra in rutas_extra:
            if isinstance(extra, dict):
                ridx = extra["idx"]
                trans_pref = extra.get("transportista")
            else:
                ridx = extra
                trans_pref = None
            grupos.append((f"R{ridx+1} · (vacía)", pd.DataFrame(columns=pedidos_df.columns), trans_pref))

    # Uniformar tuplas (por si alguna viniera de 2 elementos)
    grupos = [g if len(g) == 3 else (g[0], g[1], None) for g in grupos]

    rutas  = []

    for nombre_zona, grp_df, trans_pref in grupos:
        # Ruta vacía (agregada manualmente sin pedidos aún)
        if grp_df.empty:
            # Si hay transportista preferido, usarlo; sino el primero
            if trans_pref:
                m = transportes_df[transportes_df["nombre"] == trans_pref]
                trans_row = m.iloc[0] if not m.empty else transportes_df.iloc[0]
            else:
                trans_row = transportes_df.iloc[0]
            cap_kg    = float(trans_row["capacidad_kg"])
            tarifa    = float(trans_row["costo_hora"])
            fijo      = float(trans_row.get("costo_fijo", 0))
            ayudante  = bool(trans_row.get("ayudante", False))
            rid       = f"{nombre_zona}_empty"
            h_ruta    = horas_man.get(rid, 4.0)
            f_auto    = auto_flete(h_ruta, tarifa, fijo, ayudante)
            flete     = fletes_man.get(rid, f_auto)
            rutas.append({
                "id": rid, "zona": nombre_zona,
                "transportista": str(trans_row["nombre"]),
                "vehiculo": str(trans_row["nombre"]),
                "cap_kg": cap_kg, "kg_total": 0, "val_total": 0, "bultos_total": 0,
                "n_paradas": 0, "horas": h_ruta,
                "tarifa_hora": tarifa, "costo_fijo": fijo, "ayudante": ayudante,
                "flete": flete, "flete_auto": f_auto,
                "impacto": 0, "pct_carga": 0, "peds": [],
            })
            continue

        kg_grp     = grp_df["kilos"].sum()
        prima_zona = grp_df["zona"].iloc[0]
        # Si hay transportista preferido (ruta extra con pedidos), usarlo
        if trans_pref:
            m = transportes_df[transportes_df["nombre"] == trans_pref]
            trans_row = m.iloc[0] if not m.empty else _elegir_transportista(kg_grp, prima_zona, transportes_df)
        else:
            trans_row = _elegir_transportista(kg_grp, prima_zona, transportes_df)
        cap_kg     = float(trans_row["capacidad_kg"])
        tarifa     = float(trans_row["costo_hora"])
        fijo       = float(trans_row.get("costo_fijo", 0))
        ayudante   = bool(trans_row.get("ayudante", False))

        # Dividir en chunks si supera capacidad o MAX_PARADAS
        chunks, chunk, kg_chunk = [], [], 0.0
        for _, ped in grp_df.sort_values("cp").iterrows():
            if (kg_chunk + ped["kilos"] > cap_kg or len(chunk) >= MAX_PARADAS) and chunk:
                chunks.append(chunk)
                chunk, kg_chunk = [], 0.0
            chunk.append(ped)
            kg_chunk += ped["kilos"]
        if chunk:
            chunks.append(chunk)

        for ci, chunk_list in enumerate(chunks):
            # ID único: transportista + índice global (evita duplicados por corredor)
            rid   = f"{str(trans_row['nombre']).replace(' ', '_')}_{len(rutas)}"
            sfx   = f" ({ci+1})" if len(chunks) > 1 else ""
            peds_rows = pd.DataFrame(chunk_list)
            kg_t  = float(peds_rows["kilos"].sum())
            val_t = float(peds_rows["valor"].sum())
            bul_t = int(peds_rows["bultos"].sum())
            n_p   = len(peds_rows)

            h_auto = auto_horas(n_p)
            h_ruta = horas_man.get(rid, h_auto)
            f_auto = auto_flete(h_ruta, tarifa, fijo, ayudante)
            flete  = fletes_man.get(rid, f_auto)

            peds_data = []
            for _, p in peds_rows.iterrows():
                peds_data.append({
                    "id":        str(p.get("id", "")),
                    "cliente":   str(p.get("cliente", "")),
                    "direccion": str(p.get("direccion", "")),
                    "zona":      str(p.get("zona", "")),
                    "localidad": str(p.get("localidad", "") or p.get("zona", "")),
                    "cp":        int(p.get("cp", 0)),
                    "kilos":     float(p["kilos"]),
                    "valor":     float(p["valor"]),
                    "bultos":    int(p.get("bultos", 0)),
                    "imp_ped":   imp_ped(p["kilos"], p["valor"], kg_t, flete),
                    "flete_ped": flete * p["kilos"] / kg_t if kg_t > 0 else 0,
                })

            rutas.append({
                "id":           rid,
                "zona":         nombre_zona + sfx,
                "transportista":str(trans_row["nombre"]),
                "vehiculo":     str(trans_row["nombre"]),
                "cap_kg":       cap_kg,
                "kg_total":     kg_t,
                "val_total":    val_t,
                "bultos_total": bul_t,
                "n_paradas":    n_p,
                "horas":        h_ruta,
                "tarifa_hora":  tarifa,
                "costo_fijo":   fijo,
                "ayudante":     ayudante,
                "flete":        flete,
                "flete_auto":   f_auto,
                "impacto":      flete / val_t * 100 if val_t > 0 else 0,
                "pct_carga":    kg_t / cap_kg * 100 if cap_kg > 0 else 0,
                "peds":         peds_data,
            })

    # ── Post-proceso: aplicar asignaciones manuales (drag & drop) ──
    # asign_man = {pedido_id: ruta_id_destino}
    # Mover el pedido de su ruta actual a la ruta destino
    if asign_man:
        # Construir mapa ruta_id → índice en rutas[]
        id_a_idx = {r["id"]: i for i, r in enumerate(rutas)}

        for pid_str, dest_ruta_id in asign_man.items():
            dest_idx = id_a_idx.get(dest_ruta_id)
            if dest_idx is None:
                continue  # ruta destino no existe, ignorar

            # Buscar el pedido en sus rutas actuales y moverlo
            for src_idx, ruta in enumerate(rutas):
                if src_idx == dest_idx:
                    continue
                peds_originales = ruta["peds"]
                ped_a_mover = next((p for p in peds_originales if str(p["id"]) == pid_str), None)
                if ped_a_mover is None:
                    continue

                # Quitar de la ruta origen
                ruta["peds"] = [p for p in peds_originales if str(p["id"]) != pid_str]
                ruta["kg_total"] -= ped_a_mover["kilos"]
                ruta["val_total"] -= ped_a_mover["valor"]
                ruta["n_paradas"] = len(ruta["peds"])

                # Agregar a la ruta destino
                dest_ruta = rutas[dest_idx]
                dest_ruta["peds"].append(ped_a_mover)
                dest_ruta["kg_total"] += ped_a_mover["kilos"]
                dest_ruta["val_total"] += ped_a_mover["valor"]
                dest_ruta["n_paradas"] = len(dest_ruta["peds"])

                # Recalcular impactos y porcentajes de carga
                for r in [ruta, dest_ruta]:
                    r["pct_carga"] = r["kg_total"] / r["cap_kg"] * 100 if r["cap_kg"] > 0 else 0
                    r["impacto"] = r["flete"] / r["val_total"] * 100 if r["val_total"] > 0 else 0
                    # Recalcular Part. x Peso de cada pedido
                    for p in r["peds"]:
                        if r["kg_total"] > 0 and p["valor"] > 0:
                            p["flete_ped"] = r["flete"] * p["kilos"] / r["kg_total"]
                            p["imp_ped"]   = p["flete_ped"] / p["valor"] * 100
                break

        # Quitar rutas que quedaron vacías (el pedido único fue movido)
        rutas = [r for r in rutas if r["n_paradas"] > 0 or r.get("_keep_empty")]

    return rutas

# ══════════════════════════════════════════════════════════
# CARGA DE DATOS
# ══════════════════════════════════════════════════════════
def load_from_file(uploaded) -> Optional[pd.DataFrame]:
    try:
        if uploaded.name.lower().endswith(".csv"):
            df = pd.read_csv(uploaded)
        else:
            df = pd.read_excel(uploaded)
        return normalizar_columnas(df)
    except Exception as e:
        st.error(f"Error al leer el archivo: {e}")
        return None

def load_from_gsheet(url_or_id, creds_json=None) -> Optional[pd.DataFrame]:
    if not GSHEETS_OK:
        st.error("gspread no instalado. Corré: pip install gspread google-auth")
        return None
    try:
        if creds_json:
            info  = json.loads(creds_json)
            creds = Credentials.from_service_account_info(
                info,
                scopes=["https://spreadsheets.google.com/feeds",
                        "https://www.googleapis.com/auth/drive"]
            )
            gc = gspread.authorize(creds)
        else:
            gc = gspread.oauth()

        if "spreadsheets/d/" in url_or_id:
            sheet_id = url_or_id.split("/d/")[1].split("/")[0]
        else:
            sheet_id = url_or_id

        ws   = gc.open_by_key(sheet_id).get_worksheet(0)
        df   = pd.DataFrame(ws.get_all_records())
        return normalizar_columnas(df)
    except Exception as e:
        st.error(f"Error Google Sheets: {e}")
        return None

# ══════════════════════════════════════════════════════════
# RENDER TARJETA
# ══════════════════════════════════════════════════════════
def render_ruta_header_html(r):
    """Genera el HTML del encabezado de la ruta (sin pedidos)."""
    imp   = r["impacto"]
    pct_c = r["pct_carga"]
    cc    = carga_col(pct_c)
    ay    = '<span class="badge badge-yellow">+ Ayudante</span>' if r["ayudante"] else ""
    imp_color = "#16a34a" if imp <= 3 else "#d97706" if imp <= 5 else "#dc2626"
    imp_w   = min(imp / 8 * 100, 100)
    carga_w = min(pct_c, 100)

    merc_min = r["flete"] / 0.03 if r["flete"] > 0 else 0
    alerta = (
        '<div style="font-size:11px;color:#dc2626;margin-top:6px">'
        'Para llegar al 3% necesitás $' + f'{merc_min:,.0f}' + ' de mercadería</div>'
    ) if imp > 3 else ""

    return (
        '<div class="ruta-card">'
        '<div class="ruta-header">'
          '<div>'
            '<span class="ruta-title">' + str(r["transportista"]) + '</span>'
            '<span class="badge badge-blue" style="margin-left:8px">' + str(r["zona"]) + '</span>'
            '<span class="badge badge-gray" style="margin-left:4px">'
              + str(r["vehiculo"]).replace("Greco - ", "") + '</span>'
            + ay +
          '</div>'
          '<span class="badge ' + ic_badge(imp) + '" style="font-size:14px">'
            + f'{imp:.2f}' + '%</span>'
        '</div>'
        '<div class="stat-row">'
          '<span>' + str(r["n_paradas"]) + ' paradas</span>'
          '<span><span class="stat-val">' + f'{r["kg_total"]:.0f}' + ' kg</span> / '
            + f'{r["cap_kg"]:.0f}' + ' kg</span>'
          '<span><span class="stat-val" style="color:' + cc + '">'
            + f'{pct_c:.0f}' + '% carga</span></span>'
          '<span>' + f'{r["horas"]:.1f}' + 'h</span>'
          '<span>Merc. <span class="stat-val">$' + f'{r["val_total"]:,.0f}' + '</span></span>'
        '</div>'
        '<div class="progress-wrap">'
          '<div class="progress-fill" style="width:' + f'{carga_w:.0f}' + '%;background:' + cc + '"></div>'
        '</div>'
        '<div class="progress-wrap" style="margin-top:4px">'
          '<div class="progress-fill" style="width:' + f'{imp_w:.0f}' + '%;background:' + imp_color + '"></div>'
        '</div>'
        '<div style="display:flex;justify-content:space-between;font-size:9px;color:#9ca3af">'
          '<span>impacto</span><span>objetivo 3%</span><span>8%</span>'
        '</div>'
        + alerta +
        '</div>'
    )


def render_pedido_card_html(p):
    """HTML de una tarjeta individual de pedido — dirección grande, cliente chico."""
    cls = ic_row(p["imp_ped"])
    return (
        '<div style="padding:10px 12px;border:1px solid #e5e7eb;border-radius:8px;margin-bottom:6px;background:#fafafa">'
          '<div style="display:flex;justify-content:space-between;gap:8px">'
            '<div style="flex:1;min-width:0">'
              '<div style="font-size:15px;font-weight:600;color:#111;line-height:1.3;margin-bottom:3px">'
                + str(p["direccion"]) + '</div>'
              '<div style="font-size:11px;color:#6b7280">'
                + str(p["cliente"]) + ' · #' + str(p["id"])
                + ' · CP ' + str(p["cp"])
                + ' · ' + str(p["zona"]) + '</div>'
            '</div>'
            '<div style="text-align:right;flex-shrink:0">'
              '<div style="font-size:13px;font-weight:600">' + f'{p["kilos"]:.0f}' + ' kg · ' + str(p["bultos"]) + ' btos</div>'
              '<div style="font-size:11px;color:#6b7280">$' + f'{p["valor"]:,.0f}' + '</div>'
              '<div class="' + cls + '" style="font-size:13px;margin-top:2px">' + f'{p["imp_ped"]:.2f}' + '%</div>'
            '</div>'
          '</div>'
        '</div>'
    )


def render_ruta_card(r):
    """Mantenido por compatibilidad — versión no interactiva."""
    html = render_ruta_header_html(r)
    peds_html = "".join(render_pedido_card_html(p) for p in r["peds"])
    st.markdown(html + '<details><summary style="cursor:pointer;font-size:12px;color:#6b7280;margin:8px 0">Ver ' + str(len(r["peds"])) + ' pedidos ▾</summary>' + peds_html + '</details>', unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════
# SIDEBAR
# ══════════════════════════════════════════════════════════
def sidebar_config():
    with st.sidebar:
        st.markdown("## 🚛 Smarket TMS")
        st.markdown("---")
        st.subheader("📂 Cargar pedidos")

        fuente = st.radio("Fuente", ["Datos de muestra", "Subir archivo", "Google Sheets"])
        pedidos_df = None

        if fuente == "Datos de muestra":
            pedidos_df = SAMPLE_PEDIDOS.copy()
            st.success(f"✓ {len(pedidos_df)} pedidos de muestra")

        elif fuente == "Subir archivo":
            up = st.file_uploader(
                "CSV o Excel (.csv / .xlsx / .xls)",
                type=["csv","xlsx","xls"],
                help="Columnas requeridas: zona o localidad, kilos, valor"
            )
            if up:
                pedidos_df = load_from_file(up)
                if pedidos_df is not None:
                    st.success(f"✓ {len(pedidos_df)} pedidos cargados")

        elif fuente == "Google Sheets":
            url = st.text_input("URL o ID de la Spreadsheet")
            with st.expander("🔑 JSON Service Account"):
                creds_json = st.text_area("Pegá el JSON", height=120)
            if url and st.button("Conectar"):
                with st.spinner("Conectando..."):
                    pedidos_df = load_from_gsheet(url, creds_json or None)
                    if pedidos_df is not None:
                        st.success(f"✓ {len(pedidos_df)} pedidos")

        st.markdown("---")
        st.subheader("🚛 Transportistas")
        trans_df = st.data_editor(
            SAMPLE_TRANSPORTES,
            num_rows="dynamic",
            use_container_width=True,
            column_config={
                "nombre":       st.column_config.TextColumn("Nombre"),
                "capacidad_kg": st.column_config.NumberColumn("Cap. kg"),
                "costo_hora":   st.column_config.NumberColumn("$/hora"),
                "costo_fijo":   st.column_config.NumberColumn("Fijo $"),
                "ayudante":     st.column_config.CheckboxColumn("Ayudante"),
            },
            hide_index=True,
            key="trans_ed",
        )
        st.markdown("---")
        st.caption("Smarket TMS v2.0")
    return pedidos_df, trans_df

# ══════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════
def _render_widget_drag_and_drop(rutas):
    """Widget HTML con drag & drop — versión que funciona en Streamlit Cloud."""
    import streamlit.components.v1 as components
    import html as html_mod

    rutas_js = []
    for i, r in enumerate(rutas):
        peds_js = []
        for p in r["peds"]:
            dir_str = str(p["direccion"])
            localidad_real = (
                str(p.get("localidad", "")).strip()
                or str(p.get("zona", "")).strip()
                or ""
            )
            corrs_p = corredores_de_localidad(localidad_real)
            corr_macro = min(corrs_p, key=lambda c: CORREDORES_SIZE.get(c, 999)) if corrs_p else ""
            MACRO = {"Norte 1":"Norte","Norte 2":"Norte","Noroeste 1":"Noroeste",
                     "Noroeste 2":"Noroeste","Oeste":"Oeste","Norte-CABA-Sur":"CABA-Sur",
                     "Sur 1":"Sur","CABA":"CABA"}
            zona_label = MACRO.get(corr_macro, corr_macro)
            peds_js.append({
                "id": str(p["id"]), "dir": dir_str, "cli": str(p["cliente"]),
                "zona": zona_label, "loc": localidad_real,
                "cp": int(p["cp"]), "kg": float(p["kilos"]),
                "btos": int(p["bultos"]), "val": float(p["valor"]),
                "imp": float(p["imp_ped"]),
            })
        rutas_js.append({
            "idx": i, "ruta_id": str(r["id"]),
            "nombre": f"R{i+1}", "transp": str(r["transportista"]),
            "zona": str(r["zona"]),
            "veh": str(r["vehiculo"]).replace("Greco - ", ""),
            "cap_kg": float(r["cap_kg"]), "kg_total": float(r["kg_total"]),
            "val_total": float(r["val_total"]), "flete": float(r["flete"]),
            "impacto": float(r["impacto"]), "pct_carga": float(r["pct_carga"]),
            "n_paradas": int(r["n_paradas"]), "ayudante": bool(r["ayudante"]),
            "peds": peds_js,
        })

    rutas_json = json.dumps(rutas_js, ensure_ascii=False)
    height = max(500, 250 + 100 * max(r["n_paradas"] for r in rutas))

    html = """
<!DOCTYPE html><html><head><style>
*{box-sizing:border-box;margin:0;padding:0;font-family:-apple-system,sans-serif}
body{background:transparent}
.panel{display:grid;grid-template-columns:repeat(auto-fit,minmax(340px,1fr));gap:12px}
.ruta{border:1px solid #e5e7eb;border-radius:12px;background:#fff;overflow:hidden;display:flex;flex-direction:column;min-height:160px}
.ruta.drag-over{border:2px solid #2563eb;background:#eff6ff}
.ruta-head{padding:10px 14px;background:#f9fafb;border-bottom:1px solid #e5e7eb}
.ruta-title{font-size:14px;font-weight:600;color:#111}
.badge{display:inline-block;padding:2px 8px;border-radius:99px;font-size:10px;font-weight:600;margin-left:4px}
.b-green{background:#dcfce7;color:#166534}.b-yellow{background:#fef9c3;color:#854d0e}
.b-red{background:#fee2e2;color:#991b1b}.b-blue{background:#dbeafe;color:#1e40af}.b-gray{background:#f3f4f6;color:#374151}
.stats{display:flex;flex-wrap:wrap;gap:10px;font-size:11px;color:#6b7280;margin-top:6px}
.stats b{color:#111}
.bar{height:4px;background:#e5e7eb;border-radius:2px;margin-top:6px;overflow:hidden}
.bar>div{height:100%;border-radius:2px;transition:width .3s}
.peds-zone{flex:1;padding:8px;min-height:80px;display:flex;flex-direction:column;gap:6px}
.ped{display:flex;background:#fafafa;border:1px solid #e5e7eb;border-radius:8px;overflow:hidden;transition:box-shadow .15s,opacity .15s;cursor:grab}
.ped:hover{box-shadow:0 2px 8px rgba(0,0,0,.1);border-color:#9ca3af}
.ped:active{cursor:grabbing}
.ped.dragging{opacity:.25;box-shadow:none}
.ped-handle{flex-shrink:0;padding:10px 8px;background:#f3f4f6;color:#9ca3af;user-select:none;display:flex;align-items:center;font-size:14px;border-right:1px solid #e5e7eb;letter-spacing:-1px}
.ped:hover .ped-handle{background:#e5e7eb;color:#374151}
.ped-body{flex:1;padding:8px 10px;min-width:0}
.ped-top-row{display:flex;justify-content:space-between;align-items:flex-start;gap:8px;margin-bottom:3px}
.ped-left{flex:1;min-width:0}
.ped-dir-row{display:flex;align-items:baseline;gap:6px;flex-wrap:wrap}
.ped-dir{font-size:15px;font-weight:700;color:#111;line-height:1.3}
.ped-kg-big{font-size:14px;font-weight:700;color:#374151;white-space:nowrap;flex-shrink:0}
.ped-right-top{display:flex;flex-direction:column;align-items:flex-end;flex-shrink:0;gap:2px}
.ped-val{font-size:12px;font-weight:600;color:#374151}
.ped-loc{font-size:12px;font-weight:600;color:#374151;margin-top:2px}
.ped-cli{font-size:11px;color:#9ca3af;margin-top:3px}
.ped-zona-tag{display:inline-block;padding:1px 7px;border-radius:99px;font-size:10px;font-weight:700;text-transform:uppercase;letter-spacing:.4px;margin-right:5px}
.ped-zona-tag.sur{background:#dcfce7;color:#166534}.ped-zona-tag.norte{background:#fef9c3;color:#854d0e}
.ped-zona-tag.caba{background:#ede9fe;color:#5b21b6}.ped-zona-tag.oeste{background:#ffedd5;color:#9a3412}
.ped-zona-tag.noroeste{background:#fce7f3;color:#9d174d}.ped-zona-tag.cabasur{background:#e0f2fe;color:#0c4a6e}
.empty-msg{text-align:center;padding:20px;color:#9ca3af;font-size:12px;font-style:italic;border:2px dashed #e5e7eb;border-radius:8px}
</style></head><body>
<div id="panel" class="panel"></div>
<script>
const RUTAS=__RUTAS_JSON__;
function impColor(v){if(v<=3)return{cls:"b-green",col:"#16a34a"};if(v<=5)return{cls:"b-yellow",col:"#d97706"};return{cls:"b-red",col:"#dc2626"};}
function cargaColor(p){if(p>=75)return"#16a34a";if(p>=45)return"#d97706";return"#dc2626";}
function fmt(n){return Math.round(n).toLocaleString("es-AR");}
function cleanDir(d){return d.replace(/, Ciudad Aut[oó]noma de Buenos Aires, Argentina/gi,"").replace(/, Buenos Aires, Argentina/gi,"").replace(/, Argentina/gi,"").replace(/(, *){2,}/g,", ").replace(/^[, ]+|[, ]+$/g,"").trim();}

function render(){
 const panel=document.getElementById("panel");panel.innerHTML="";
 RUTAS.forEach((r,ri)=>{
  const imp=impColor(r.impacto),cc=cargaColor(r.pct_carga);
  const ay=r.ayudante?'<span class="badge b-yellow">+ Ay.</span>':'';
  const ruta=document.createElement("div");ruta.className="ruta";ruta.dataset.idx=ri;
  ruta.innerHTML=`
   <div class="ruta-head">
    <div style="display:flex;justify-content:space-between;align-items:center;flex-wrap:wrap;gap:4px">
     <div><span class="ruta-title">${r.transp}</span><span class="badge b-blue">${r.zona}</span><span class="badge b-gray">${r.veh}</span>${ay}</div>
     <span class="badge ${imp.cls}" style="font-size:13px">${r.impacto.toFixed(2)}%</span>
    </div>
    <div class="stats">
     <span>${r.n_paradas} par.</span><span><b>${fmt(r.kg_total)} kg</b> / ${fmt(r.cap_kg)} kg</span>
     <span style="color:${cc};font-weight:600">${r.pct_carga.toFixed(0)}% carga</span>
     <span>Flete <b>$${fmt(r.flete)}</b></span><span>Merc. <b>$${fmt(r.val_total)}</b></span>
    </div>
    <div class="bar"><div style="width:${Math.min(r.pct_carga,100)}%;background:${cc}"></div></div>
    <div class="bar" style="margin-top:3px"><div style="width:${Math.min(r.impacto/8*100,100)}%;background:${imp.col}"></div></div>
   </div>
   <div class="peds-zone" id="zone-${ri}" data-idx="${ri}"></div>`;
  const zone=ruta.querySelector(".peds-zone");
  if(!r.peds.length){zone.innerHTML='<div class="empty-msg">Soltá pedidos acá</div>';}
  else{r.peds.forEach(p=>{
   const im=impColor(p.imp);
   const zonaCls={"Sur":"sur","Norte":"norte","CABA":"caba","Oeste":"oeste","Noroeste":"noroeste","CABA-Sur":"cabasur"}[p.zona]||"";
   const dirClean=cleanDir(p.dir);
   const ped=document.createElement("div");ped.className="ped";ped.draggable=true;
   ped.innerHTML=`
    <div class="ped-handle">⋮⋮</div>
    <div class="ped-body">
     <div class="ped-top-row">
      <div class="ped-left">
       <div class="ped-dir-row"><span class="ped-dir">${dirClean}</span><span class="ped-kg-big">${p.kg.toFixed(0)} kg</span></div>
       <div class="ped-loc"><span class="ped-zona-tag ${zonaCls}">${p.zona}</span>${p.loc||p.zona}</div>
      </div>
      <div class="ped-right-top"><div class="ped-val">$${fmt(p.val)}</div><div style="font-size:13px;font-weight:700;color:${im.col}">${p.imp.toFixed(2)}%</div></div>
     </div>
     <div class="ped-cli">${p.cli} · #${p.id} · CP ${p.cp} · ${p.btos} btos</div>
    </div>`;
   ped.addEventListener("dragstart",(e)=>{e.dataTransfer.setData("pid",p.id);e.dataTransfer.setData("from",String(ri));ped.classList.add("dragging");});
   ped.addEventListener("dragend",()=>{ped.classList.remove("dragging");});
   zone.appendChild(ped);
  });}
  ruta.addEventListener("dragover",(e)=>{e.preventDefault();ruta.classList.add("drag-over");});
  ruta.addEventListener("dragleave",()=>{ruta.classList.remove("drag-over");});
  ruta.addEventListener("drop",(e)=>{
   e.preventDefault();ruta.classList.remove("drag-over");
   const pid=e.dataTransfer.getData("pid"),fromIdx=parseInt(e.dataTransfer.getData("from")),toIdx=ri;
   if(fromIdx===toIdx||!pid)return;
   const url=new URL(window.parent.location.href);
   url.searchParams.set("move_pid",pid);url.searchParams.set("move_to_id",r.ruta_id);
   url.searchParams.set("_t",String(Date.now()));
   window.parent.location.href=url.toString();
  });
  panel.appendChild(ruta);
 });
}
render();
</script></body></html>
"""
    html = html.replace("__RUTAS_JSON__", rutas_json)
    components.html(html, height=height, scrolling=True)


def main():
    # Estado de la sesión
    if "fletes_man"   not in st.session_state: st.session_state["fletes_man"]   = {}
    if "horas_man"    not in st.session_state: st.session_state["horas_man"]    = {}
    if "asign_man"    not in st.session_state: st.session_state["asign_man"]    = {}
    if "rutas_extra"  not in st.session_state: st.session_state["rutas_extra"]  = []

    # Capturar query params del widget drag & drop
    qp = st.query_params
    if "move_pid" in qp and "move_to_id" in qp:
        try:
            pid     = str(qp["move_pid"])
            ruta_id = str(qp["move_to_id"])
            st.session_state["asign_man"][pid] = ruta_id
        except Exception:
            pass
        st.query_params.clear()

    pedidos_df, trans_df = sidebar_config()

    st.title("🚛 Smarket TMS — Optimizador de Ruteo")

    if pedidos_df is None:
        st.info("← Cargá pedidos desde la barra lateral para comenzar.")
        return

    # Optimizar con asignaciones manuales + rutas extra
    rutas = optimizar_rutas(
        pedidos_df, trans_df,
        st.session_state["fletes_man"],
        st.session_state["horas_man"],
        asign_man=st.session_state["asign_man"],
        rutas_extra=st.session_state["rutas_extra"],
    )

    tab_panel, tab_peds, tab_cfg, tab_exp, tab_gs = st.tabs([
        "📋 Panel de Rutas", "📦 Pedidos",
        "⚙️ Fletes", "⬇️ Exportar Excel", "🔗 Google Sheets",
    ])

        # ── PANEL INTERACTIVO (widget HTML con drag & drop) ──
    with tab_panel:
        # ── PASO 1: Capturar ediciones de flete ANTES de mostrar el widget ──
        # Esto garantiza que cuando el usuario edita un flete,
        # el widget de arriba se rerenderiza con el valor actualizado.
        if rutas:
            st.markdown("#### 💰 Editar flete por ruta")
            cols_flete = st.columns(max(1, len(rutas)))
            for i, r in enumerate(rutas):
                with cols_flete[i % len(cols_flete)]:
                    fn = st.number_input(
                        f"R{i+1} · {r['transportista']} ({r['zona'][:20]})",
                        min_value=0,
                        value=int(r["flete"]),
                        step=1000,
                        key=f"fl_p_{r['id']}_{i}",
                        help=f"Auto: ${r['flete_auto']:,.0f}",
                    )
                    # Persistir override en session_state
                    if fn != int(r["flete_auto"]):
                        st.session_state["fletes_man"][r["id"]] = fn
                    elif r["id"] in st.session_state["fletes_man"]:
                        del st.session_state["fletes_man"][r["id"]]
                    # Actualizar el objeto ruta en memoria (para que el widget lo vea)
                    r["flete"] = fn
                    r["impacto"] = fn / r["val_total"] * 100 if r["val_total"] > 0 else 0
                    # Recalcular impacto por pedido
                    for p in r["peds"]:
                        if r["kg_total"] > 0:
                            p["flete_ped"] = fn * p["kilos"] / r["kg_total"]
                            p["imp_ped"]   = (fn * p["kilos"] / r["kg_total"]) / p["valor"] * 100 if p["valor"] > 0 else 0

                    ip = r["impacto"]
                    color = "#16a34a" if ip <= 3 else "#d97706" if ip <= 5 else "#dc2626"
                    st.markdown(
                        f'<div style="margin-top:-8px;font-size:13px;color:{color};font-weight:600">Impacto: {ip:.2f}%</div>',
                        unsafe_allow_html=True,
                    )

            st.markdown("---")

        # ── PASO 2: KPIs con los valores ya actualizados ─────
        tc  = sum(r["flete"]     for r in rutas)
        tv  = sum(r["val_total"] for r in rutas)
        tkg = sum(r["kg_total"]  for r in rutas)
        ig  = tc / tv * 100 if tv > 0 else 0

        c1,c2,c3,c4,c5 = st.columns(5)
        c1.metric("Rutas activas", len(rutas))
        c2.metric("Pedidos",       sum(r["n_paradas"] for r in rutas))
        c3.metric("Kg total",      f"{tkg:,.0f}")
        c4.metric("Costo flete",   f"${tc:,.0f}")
        c5.metric("Impacto global",f"{ig:.2f}%",
                  delta=f"{ig-3:.2f}% vs 3%", delta_color="inverse")

        # ── PASO 3: Acciones globales (agregar flete con selector de transportista) ──
        st.markdown("---")
        st.markdown("#### ➕ Agregar un nuevo flete manual")
        # Transportistas únicos ya asignados en las rutas actuales
        usados = set(r["transportista"] for r in rutas)
        disponibles = []
        for nom in trans_df["nombre"].tolist():
            # Excluir JG únicos que ya estén en uso
            if nom in TRANSPORTISTAS_UNICOS and nom in usados:
                continue
            disponibles.append(nom)

        col_sel, col_btn, col_reset = st.columns([3, 1, 1])
        with col_sel:
            if disponibles:
                trans_nuevo = st.selectbox(
                    "Elegí transportista para el nuevo flete",
                    options=disponibles,
                    key="sel_nuevo_trans",
                    label_visibility="collapsed",
                    help="Gaby/Juan/Ezequiel solo pueden hacer 1 recorrido por día. Greco puede hacer varios.",
                )
            else:
                trans_nuevo = None
                st.info("No hay transportistas disponibles.")
        with col_btn:
            disabled = trans_nuevo is None
            if st.button("➕ Agregar", use_container_width=True, type="primary", disabled=disabled):
                nuevo_idx = max([len(rutas)] + [
                    (e["idx"] if isinstance(e, dict) else e) + 1
                    for e in st.session_state["rutas_extra"]
                ] + [0])
                st.session_state["rutas_extra"].append({
                    "idx": nuevo_idx,
                    "transportista": trans_nuevo,
                })
                st.rerun()
        with col_reset:
            if st.button("🔄 Resetear", use_container_width=True,
                         help="Vuelve a la asignación automática"):
                st.session_state["asign_man"]   = {}
                st.session_state["rutas_extra"] = []
                st.session_state["fletes_man"]  = {}
                st.rerun()

        st.markdown("---")

        # ── PASO 4: Widget con drag & drop (ya con fletes actualizados) ──
        if rutas:
            criticas = [r for r in rutas if r["impacto"] > 3]
            if criticas:
                st.warning(f"⚠️ {len(criticas)} ruta(s) superan el 3% de impacto.")
            else:
                st.success("✅ Todas las rutas están por debajo del 3%.")

            _render_widget_drag_and_drop(rutas)
        else:
            st.info("No hay rutas generadas.")

    # ── PEDIDOS ───────────────────────────────────────────
    with tab_peds:
        st.subheader("Pedidos cargados")
        st.dataframe(pedidos_df, use_container_width=True, hide_index=True)
        st.caption(f"{len(pedidos_df)} pedidos · {pedidos_df['kilos'].sum():,.0f} kg · ${pedidos_df['valor'].sum():,.0f}")

    # ── FLETES ────────────────────────────────────────────
    with tab_cfg:
        st.subheader("⚙️ Editar flete y horas por ruta")
        if not rutas:
            st.info("Cargá pedidos para ver las rutas.")
        else:
            for r in rutas:
                with st.expander(f"**{r['transportista']}** — {r['zona']}  |  Impacto: {r['impacto']:.2f}%"):
                    col1, col2, col3 = st.columns(3)
                    with col1:
                        fn = st.number_input(
                            "Flete total ($)", min_value=0, value=int(r["flete"]),
                            step=1000, key=f"fl_{r['id']}",
                            help=f"Auto: ${r['flete_auto']:,.0f}"
                        )
                    with col2:
                        hn = st.number_input(
                            "Horas", min_value=0.5, max_value=24.0,
                            value=float(r["horas"]), step=0.5, key=f"hr_{r['id']}"
                        )
                    with col3:
                        ip = fn / r["val_total"] * 100 if r["val_total"] > 0 else 0
                        color = "green" if ip <= 3 else "orange" if ip <= 5 else "red"
                        st.markdown("**Preview impacto:**")
                        st.markdown(f"<h2 style=\'color:{color}\'>{ip:.2f}%</h2>", unsafe_allow_html=True)

                    if fn != int(r["flete_auto"]):
                        st.session_state["fletes_man"][r["id"]] = fn
                    elif r["id"] in st.session_state["fletes_man"]:
                        del st.session_state["fletes_man"][r["id"]]
                    st.session_state["horas_man"][r["id"]] = hn

                    mm = fn / 0.03 if fn > 0 else 0
                    st.caption(f"Para 3%: necesitás ${mm:,.0f} (tenés ${r['val_total']:,.0f})")

    # ── EXPORTAR ──────────────────────────────────────────
    with tab_exp:
        st.subheader("⬇️ Exportar Hojas de Ruta")
        if not rutas:
            st.info("Cargá pedidos para exportar.")
        elif st.button("📊 Generar Excel", type="primary", use_container_width=True):
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                from openpyxl.utils import get_column_letter

                def fill(h): return PatternFill("solid", fgColor=h)
                def th():
                    s = Side(style="thin", color="CCCCCC")
                    return Border(left=s, right=s, top=s, bottom=s)
                def fic(v): return "EAF3DE" if v<=3 else "FAEEDA" if v<=5 else "FCEBEB"

                wb = Workbook()
                wb.remove(wb.active)

                # Resumen
                wsr = wb.create_sheet("Resumen", 0)
                for i, w in enumerate([14,22,16,8,8,10,14,12,18], 1):
                    wsr.column_dimensions[get_column_letter(i)].width = w
                wsr.merge_cells("A1:I1")
                wsr["A1"] = "SMARKET — RESUMEN HOJAS DE RUTA"
                wsr["A1"].font = Font(name="Arial", size=12, bold=True, color="FFFFFF")
                wsr["A1"].fill = fill("1F3864")
                wsr["A1"].alignment = Alignment(horizontal="center")

                for ci, h in enumerate(["Transportista","Zona","Vehículo","Horas","Kg","Paradas","Flete","Impacto %","Valor"],1):
                    c = wsr.cell(2,ci,h)
                    c.font = Font(name="Arial",size=9,bold=True,color="FFFFFF")
                    c.fill = fill("2E75B6"); c.border = th()
                    c.alignment = Alignment(horizontal="center")

                for ri, r in enumerate(rutas):
                    row = 3+ri
                    bg = "E1F5EE" if r["transportista"] in ("Gaby","Juan") else "EEEDFE"
                    for ci, v in enumerate([r["transportista"],r["zona"],r["vehiculo"],
                                            r["horas"],f'{r["kg_total"]:.0f}',r["n_paradas"],
                                            r["flete"],f'{r["impacto"]:.2f}%',r["val_total"]],1):
                        c = wsr.cell(row,ci,v)
                        c.font = Font(name="Arial",size=9)
                        c.fill = fill(fic(r["impacto"]) if ci==8 else bg)
                        c.border = th()
                        c.alignment = Alignment(horizontal="center" if ci>2 else "left")
                        if ci in (7,9): c.number_format = "$#,##0"

                # Una hoja por ruta
                for r in rutas:
                    sname = f"{r['transportista'][:10]}-{r['zona'][:12]}"
                    ws2 = wb.create_sheet(sname)
                    for i,w in enumerate([5,10,28,38,16,7,7,7,12,8,22,13,13],1):
                        ws2.column_dimensions[get_column_letter(i)].width = w
                    ws2.merge_cells("A1:M1")
                    ws2["A1"] = f"SMARKET — {r['transportista']} — {r['zona']}"
                    ws2["A1"].font = Font(name="Arial",size=11,bold=True,color="FFFFFF")
                    ws2["A1"].fill = fill("1F3864")
                    ws2["A1"].alignment = Alignment(horizontal="center")
                    ay_s = " + Ayudante" if r["ayudante"] else ""
                    ws2.merge_cells("A2:M2")
                    ws2["A2"] = f"Transp: {r['transportista']} | Veh: {r['vehiculo']}{ay_s} | Horas: {r['horas']}h | Flete: ${r['flete']:,.0f} | Impacto: {r['impacto']:.2f}%"
                    ws2["A2"].font = Font(name="Arial",size=8,bold=True)
                    ws2["A2"].fill = fill("E1F5EE" if r["transportista"] in ("Gaby","Juan") else "EEEDFE")
                    for ci, h in enumerate(["Ord.","N° Ped.","Cliente","Dirección","Zona","CP","Kg","Btos","Valor","Horario","Obs","Flete$","Part.%"],1):
                        c = ws2.cell(4,ci,h)
                        c.font = Font(name="Arial",size=8,bold=True,color="FFFFFF")
                        c.fill = fill("2E75B6"); c.border = th()
                        c.alignment = Alignment(horizontal="center",wrap_text=True)
                    for ni, p in enumerate(r["peds"]):
                        row = 5+ni
                        bg2 = "FFFFFF" if ni%2==0 else "EBF3FB"
                        for ci, v in enumerate([ni+1,p["id"],p["cliente"],p["direccion"],p["zona"],
                                                p["cp"],p["kilos"],p["bultos"],p["valor"],
                                                "","",p["flete_ped"],f'{p["imp_ped"]:.2f}%'],1):
                            c = ws2.cell(row,ci,v)
                            c.font = Font(name="Arial",size=8)
                            c.fill = fill(fic(p["imp_ped"]) if ci==13 else bg2)
                            c.border = th()
                            c.alignment = Alignment(horizontal="left" if ci in(3,4,11) else "center",wrap_text=ci in(3,4,11))
                            if ci in (9,12): c.number_format = "$#,##0"
                        ws2.row_dimensions[row].height = 28
                    ws2.freeze_panes = "A5"
                    ws2.page_setup.orientation = "landscape"
                    ws2.page_setup.fitToPage = True

                buf = io.BytesIO()
                wb.save(buf)
                buf.seek(0)
                st.download_button(
                    "⬇️ Descargar Smarket_HDR.xlsx", buf,
                    file_name="Smarket_HDR.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )
                st.success("✅ Excel listo. Hacé click en Descargar.")
            except Exception as e:
                st.error(f"Error generando Excel: {e}")

    # ── GOOGLE SHEETS ──────────────────────────────────────
    with tab_gs:
        st.subheader("🔗 Configurar Google Sheets")
        st.markdown("""
### Pasos para conectar

**1. Google Cloud Console**
- Creá un proyecto en [console.cloud.google.com](https://console.cloud.google.com)
- Activá **Google Sheets API** y **Google Drive API**

**2. Service Account**
- IAM y Admin → Service Accounts → Crear
- Rol: Editor (o Viewer si solo vas a leer)
- Crear clave → formato JSON → descargar

**3. Compartir la Spreadsheet**
- Abrí tu Google Sheet → Compartir
- Pegá el email de la Service Account (`xxx@proyecto.iam.gserviceaccount.com`)
- Permisos de Lector

**4. Formato esperado de la hoja**

| N° de Pedido | Descripción Cliente | Dirección de entrega | Localidad | CP | Kilos | Valor | Bultos |
|---|---|---|---|---|---|---|---|
| 35268 | Sebastian Duarte | Av. S. Martín 7300 | Villa Devoto | 1419 | 114 | 570137 | 74 |

**5. En producción (secrets.toml)**
```toml
[gsheets]
type = "service_account"
project_id = "tu-proyecto"
private_key = "-----BEGIN PRIVATE KEY-----\n..."
client_email = "smarket@tu-proyecto.iam.gserviceaccount.com"
```
        """)


if __name__ == "__main__":
    main()
